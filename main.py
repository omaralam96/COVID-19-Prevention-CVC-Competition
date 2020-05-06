# -*- coding: utf-8 -*-
"""
Created on Sun May  3 16:55:07 2020
@author: Mark
"""
from PyQt5 import uic
from PyQt5.QtWidgets import QApplication,QFileDialog
############
import numpy as np
from PyQt5 import QtCore, QtGui
import qimage2ndarray
import cv2
from camera_algorithms import camera1
from camera_algorithms import camera2
import os
import sys
from os import path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

flag_play=0
video_source=""
video_path=""
workbook_name = "Attendance.xlsx"
#Creating excel sheet for attendance
if path.exists(workbook_name):
    wb =  load_workbook(filename = workbook_name)
    sheet = wb.active                
else:
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "ID"
    sheet["B1"] = "Time"
    sheet["C1"] = "Date"  
    wb.save(filename=workbook_name)                        

def Program():
    global form,cap,flag_play,video_source,wb,sheet
    index=form.tabWidget.currentIndex()
    #In case we are in tab1(Social distancing) of GUI and user click play button
    if index == 0 and flag_play == 1:      
        #Loading the frame from selected video
        success, img_before = video_source.read() 
        #Checking video finsihes or not
        if np.shape(img_before) != (): 
            #Converting img_before from np.array to suitable format for GUI
            img=cv2.cvtColor(img_before, cv2.COLOR_BGR2RGB)
            qImg = qimage2ndarray.array2qimage(img)
            form.video_before.setPixmap(QtGui.QPixmap(qImg))
            #CALLING Social Distancing System 
            img_after,n_total,n_malPositioned=camera1.queue(img_before)
            #Converting img_After from np.array to suitable format for GUI            
            img=cv2.cvtColor(img_after, cv2.COLOR_BGR2RGB)
            qImg = qimage2ndarray.array2qimage(img)
            #Updating GUI Interface 
            form.video_after.setPixmap(QtGui.QPixmap(qImg))
            form.queue_total.setText(str(n_total))
            form.queue_danger.setText(str(n_malPositioned))
        else:
            #reset play_flag of play button when playing video finishes 
            flag_play=0
            finish_pic = cv2.imread("images/handbrake.png")
            img=cv2.cvtColor(finish_pic, cv2.COLOR_BGR2RGB)
            qImg = qimage2ndarray.array2qimage(img)
            #Updating GUI Interface 
            form.video_before.setPixmap(QtGui.QPixmap(qImg))
            form.video_after.setPixmap(QtGui.QPixmap(qImg))
    #In case we are in tab2(Login system) of GUI    
    elif index == 1:
        # Read the frame from camera
        _, img_before = cap.read()
        ##CALLING Login system 
        img_after,mask_label,glove_label,emp_Id,emp_Name,emp_Time,profile_pic = camera2.loginSystem(img_before)
        #Converting img_After from np.array to suitable format for GUI            
        img = cv2.cvtColor(img_after, cv2.COLOR_BGR2RGB)        
        qImg = qimage2ndarray.array2qimage(img)
        #UPDATING GUI VARIABLES
        form.cam2view.setPixmap(QtGui.QPixmap(qImg))
        form.maskstatus.setText(mask_label)
        form.glovestatus.setText(glove_label)
        form.emp_id.setText(emp_Id)
        form.emp_name.setText(emp_Name)
        form.emp_time.setText(emp_Time)
        #Update the profile picture of employee if exist
        if profile_pic !="" :
            img = cv2.cvtColor(profile_pic, cv2.COLOR_BGR2RGB)
            qImg = qimage2ndarray.array2qimage(img)
            form.emp_profile.setPixmap(QtGui.QPixmap(qImg))

#Event handler of browse button
def browse_handler():
    global video_source,flag_play,video_path
    #reset the play_flag of play button if the user want to select another file
    flag_play = 0
    qfd = QFileDialog()
    #Get path of selected file
    filename = QFileDialog.getOpenFileName(qfd, filter = "mp4(*.mp4)")   
    video_path=filename[0] 
    if(video_path != ""):
        form.video_name.setText(path.basename(video_path))
        #Enable play button and Display the first frame of Video 
        form.play.setEnabled(True)
        video_source=cv2.VideoCapture(video_path)
        success, first_frame = video_source.read()
        #Converting first_frame from np.array to suitable format for GUI            
        img=cv2.cvtColor(first_frame, cv2.COLOR_BGR2RGB)
        qImg = qimage2ndarray.array2qimage(img)
        #UPDATING GUI VARIABLES
        form.video_before.setPixmap(QtGui.QPixmap(qImg))
    else:
        form.video_name.setText("no file selected")
     
#Event handler of play button
def play_handler():
    global flag_play,video_source,video_path
    video_source=""
    #Loading the video using video path
    video_source=cv2.VideoCapture(video_path)
    #set play flag
    flag_play=1
    
#Event handler of attendance sheet button
def sheet_handler(): 
    #Get the path of attendance sheet
    pathname = os.path.dirname(sys.argv[0])        
    #Open the attendance sheet
    os.system('start excel.exe "%s\\Attendance.xlsx"' % (os.path.abspath(pathname), ))
   
    
#Loading GUI 
Form, Window = uic.loadUiType("gui.ui")  
app = QApplication([])
window = Window()
form = Form()                               
form.setupUi(window)
#
cap = cv2.VideoCapture(0)
##Creating Timer that Updates the GUI and Operates the program that contains social distancing system and login system   
timer = QtCore.QTimer()
timer.timeout.connect(Program)
timer.start(60)
#Add event handlers to browse,play and attendance sheet buttons
form.browse.clicked.connect(browse_handler)
form.play.clicked.connect(play_handler)
form.sheet_button.clicked.connect(sheet_handler)
#disable play button until the user select video
form.play.setEnabled(False)
window.show()
app.exec_()